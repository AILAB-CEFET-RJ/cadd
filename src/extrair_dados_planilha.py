import sys, getopt
import unicodedata
from openpyxl import load_workbook


def dadosPlanilha(filename):
    wb = load_workbook(filename=filename, read_only=True)
    # Pega a primeira aba (ou ajuste para nome fixo se necessário)
    ws = wb.active  
    
    dados_agregados = {}
    qtd_matriculas = 0
    
    print("Total de linhas = %d" % ws.max_row)
    
    # Itera pelas linhas a partir da segunda (ignorando cabeçalho)
    for rowNum, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if rowNum % 500 == 0:
            print("# linhas processadas %d" % rowNum)
        
        ano = row[9]   # coluna 10 (0-indexed → 9)
        situacao = row[11]  # coluna 12 (0-indexed → 11)

        if ano is None or situacao is None:
            continue

        tupla = dados_agregados.get(ano, (0, 0, 0))

        # Normaliza situação
        situacao = unicodedata.normalize('NFKD', str(situacao)).encode('ascii', 'ignore').decode('ascii')

        if situacao.startswith('Aprovado') or situacao.startswith('Isento') or situacao.startswith('Aproveitamento'):
            tupla = (tupla[0] + 1, tupla[1], tupla[2])
        elif situacao.startswith('Reprovado'):
            tupla = (tupla[0], tupla[1] + 1, tupla[2])
        elif situacao.startswith('Trancamento'):
            tupla = (tupla[0], tupla[1], tupla[2] + 1)
        elif situacao.startswith('Matricula'):
            qtd_matriculas += 1
        else:
            print("Situacao ignorada: %s" % situacao)

        dados_agregados[ano] = tupla
    
    print("# matriculas = %d" % qtd_matriculas)
    
    for key, value in dados_agregados.items():
        print("(%s) %s" % (key, value))


def main(argv):
    filename = None

    try:
        opts, args = getopt.getopt(argv, "hf:", ["file="])
    except getopt.GetoptError:
        print("main.py -f <excel_file>")
        sys.exit(2)

    for opt, arg in opts:
        if opt == '-h':
            print("main.py -f <excel_file>")
            sys.exit()
        elif opt in ("-f", "--file"):
            filename = arg

    if not filename:
        print("Erro: é necessário informar o arquivo Excel via -f")
        sys.exit(2)

    dadosPlanilha(filename)


if __name__ == "__main__":
    main(sys.argv[1:])

# Como usar:
# python main.py -f "11.02.05.99.60-GPROD.xlsx"
